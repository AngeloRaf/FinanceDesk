"""
FinanceDesk v1.1 — export_service.py
Export Excel (openpyxl) et PDF (reportlab) pour tous les modules.
ReportLab importé en lazy pour éviter l'erreur PIL au démarrage.
"""

import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE = "1A3A6B"
GRAY = "F4F4F0"


def export_excel(data, colonnes, titres, nom_feuille, titre_rapport, date_debut="", date_fin=""):
    wb = Workbook()
    ws = wb.active
    ws.title = nom_feuille
    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor=BLUE)
    hdr_align = Alignment(horizontal="center", vertical="center")
    title_font= Font(name="Calibri", bold=True, size=14, color=BLUE)
    sub_font  = Font(name="Calibri", size=10, color="5A5A55")
    thin      = Side(style="thin", color="E5E5E0")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(f"A1:{chr(64+len(colonnes))}1")
    ws["A1"] = titre_rapport
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    if date_debut or date_fin:
        periode = f"Période : {date_debut or '—'}  →  {date_fin or '—'}"
        ws.merge_cells(f"A2:{chr(64+len(colonnes))}2")
        ws["A2"] = periode
        ws["A2"].font = sub_font
        ws.row_dimensions[2].height = 18

    row_hdr = 3
    for col_idx, titre in enumerate(titres, 1):
        cell = ws.cell(row=row_hdr, column=col_idx, value=titre)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
    ws.row_dimensions[row_hdr].height = 22

    for row_idx, item in enumerate(data, row_hdr + 1):
        fill = PatternFill("solid", fgColor="FFFFFF" if row_idx % 2 == 0 else GRAY)
        for col_idx, key in enumerate(colonnes, 1):
            val  = item.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill; cell.border = border
            cell.alignment = Alignment(vertical="center")
            if isinstance(val, float):
                cell.number_format = "#,##0.00 €"

    for col_idx, key in enumerate(colonnes, 1):
        col_letter = chr(64 + col_idx)
        max_len = max(len(str(titres[col_idx-1])),
                      max((len(str(r.get(key,""))) for r in data), default=0))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    path = _get_export_path(nom_feuille, "xlsx")
    wb.save(path)
    return str(path)


def export_factures_excel(factures, date_debut="", date_fin=""):
    return export_excel(factures,
        ["numero","fournisseur","montant","date_echeance","statut","date_paiement","mode_reglement","commentaire"],
        ["N° Facture","Fournisseur","Montant","Échéance","Statut","Date paiement","Règlement","Commentaire"],
        "Factures", "Rapport Factures — FinanceDesk", date_debut, date_fin)

def export_virements_excel(virements, date_debut="", date_fin=""):
    return export_excel(virements,
        ["date_virement","beneficiaire","montant","ref_transaction","commentaire"],
        ["Date","Bénéficiaire","Montant","Référence","Commentaire"],
        "Virements", "Rapport Virements — FinanceDesk", date_debut, date_fin)

def export_caisse_excel(operations, date_debut="", date_fin=""):
    return export_excel(operations,
        ["date_operation","description","categorie","type_operation","montant","solde_apres","justificatif"],
        ["Date","Description","Catégorie","Type","Montant","Solde après","Justificatif"],
        "Petite Caisse", "Rapport Petite Caisse — FinanceDesk", date_debut, date_fin)

def export_recettes_excel(recettes, date_debut="", date_fin=""):
    return export_excel(recettes,
        ["date_reception","nom_payeur","numero_facture","ref_transaction","montant"],
        ["Date","Payeur","N° Facture","Référence","Montant"],
        "Recettes", "Rapport Recettes — FinanceDesk", date_debut, date_fin)


def export_rapport_pdf(data_rapport, date_debut, date_fin):
    # Imports lazy — chargés uniquement quand on génère un PDF
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    path = _get_export_path("rapport_financier", "pdf")
    doc  = SimpleDocTemplate(str(path), pagesize=A4,
                              leftMargin=2*cm, rightMargin=2*cm,
                              topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    title_style = ParagraphStyle("FDTitle", parent=styles["Title"],
        fontSize=18, textColor=colors.HexColor(f"#{BLUE}"), spaceAfter=6)
    h2_style = ParagraphStyle("FDH2", parent=styles["Heading2"],
        fontSize=13, textColor=colors.HexColor(f"#{BLUE}"), spaceBefore=16, spaceAfter=6)

    story.append(Paragraph("FinanceDesk — Rapport Financier", title_style))
    story.append(Paragraph(f"Période : {date_debut} → {date_fin}", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))

    stats = data_rapport.get("stats", {})
    kpi_data = [
        ["Indicateur", "Valeur"],
        ["Total factures à payer",  f"{stats.get('total_a_payer', 0):,.2f} €"],
        ["Total virements période", f"{stats.get('total_virements', 0):,.2f} €"],
        ["Total recettes période",  f"{stats.get('total_recettes', 0):,.2f} €"],
        ["Solde petite caisse",     f"{stats.get('solde_caisse', 0):,.2f} €"],
        ["Solde net",               f"{stats.get('solde_net', 0):,.2f} €"],
    ]
    story.append(Paragraph("Résumé", h2_style))
    story.append(_pdf_table(kpi_data, colors, col_widths=[10*cm, 5*cm]))
    story.append(Spacer(1, 0.4*cm))

    for titre_s, rows, cols in [
        ("Factures",     data_rapport.get("factures",[]), ["numero","fournisseur","montant","date_echeance","statut"]),
        ("Virements",    data_rapport.get("virements",[]),["date_virement","beneficiaire","montant","ref_transaction"]),
        ("Petite Caisse",data_rapport.get("caisse",[]),   ["date_operation","description","type_operation","montant"]),
        ("Recettes",     data_rapport.get("recettes",[]), ["date_reception","nom_payeur","montant","ref_transaction"]),
    ]:
        if not rows: continue
        story.append(Paragraph(titre_s, h2_style))
        tdata = [cols] + [[str(r.get(c,"")) for c in cols] for r in rows]
        story.append(_pdf_table(tdata, colors))
        story.append(Spacer(1, 0.3*cm))

    doc.build(story)
    return str(path)


def _pdf_table(data, colors, col_widths=None):
    from reportlab.platypus import Table, TableStyle
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",     (0,0),(-1,0), colors.HexColor(f"#{BLUE}")),
        ("TEXTCOLOR",      (0,0),(-1,0), colors.white),
        ("FONTNAME",       (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",       (0,0),(-1,-1), 9),
        ("ALIGN",          (0,0),(-1,-1), "LEFT"),
        ("VALIGN",         (0,0),(-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1),(-1,-1), [colors.white, colors.HexColor(f"#{GRAY}")]),
        ("GRID",           (0,0),(-1,-1), 0.3, colors.HexColor("#E5E5E0")),
        ("TOPPADDING",     (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",  (0,0),(-1,-1), 4),
    ]))
    return t


def _get_export_path(nom, ext):
    from core.db_manager import BASE_DIR
    export_dir = BASE_DIR / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return export_dir / f"{nom}_{ts}.{ext}"